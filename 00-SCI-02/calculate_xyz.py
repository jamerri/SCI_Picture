# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     calculate_xyz
   Description :
   Author :       Jamerri
   date：          2022/10/29
-------------------------------------------------
   Change Activity:
                   2022/10/29:
-------------------------------------------------
"""


import xlrd
from openpyxl.workbook import Workbook  # 新建文件用

# 读取路径
book = xlrd.open_workbook(filename=r"IPSO-3D-INV-135.xlsx")

# 读取名字为Sheet1的表
sheet1 = book.sheet_by_name("1_robot")

# 读取名字为Sheet2的表
sheet2 = book.sheet_by_name("2_robot")

# 读取名字为Sheet3的表
sheet3 = book.sheet_by_name("3_robot")


# 判断此次实验的步数
rowamount = sheet1.nrows
step_total = rowamount - 12
print("step total is ", step_total)

# 用于存储1号机器人的数组
robot_1_x = []
robot_1_y = []
robot_1_z = []

row_num = 8
while row_num <= (step_total + 8):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_1_x.append(4.97 - float(sheet1.cell_value(row_num, 2)))
    robot_1_y.append(2.38 + float(sheet1.cell_value(row_num, 1)))
    robot_1_z.append(float(sheet1.cell_value(row_num, 3)))
    row_num = row_num + 1
# # print(robot_1_x)
# print("This experimental points are " + str(len(robot_1_x)) + "!!!!")

# 用于存储2号机器人的数组
robot_2_x = []
robot_2_y = []
robot_2_z = []
row_num = 8

while row_num <= (step_total + 8):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_2_x.append(4.97 - float(sheet2.cell_value(row_num, 2)))
    robot_2_y.append(2.38 + float(sheet2.cell_value(row_num, 1)))
    robot_2_z.append(float(sheet2.cell_value(row_num, 3)))
    row_num = row_num + 1

# 用于存储3号机器人的数组
robot_3_x = []
robot_3_y = []
robot_3_z = []
row_num = 8

while row_num <= (step_total + 8):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_3_x.append(4.97 - float(sheet3.cell_value(row_num, 2)))
    robot_3_y.append(2.38 + float(sheet3.cell_value(row_num, 1)))
    robot_3_z.append(float(sheet3.cell_value(row_num, 3)))
    row_num = row_num + 1

# print(robot_1_x)
# print("len is ", len(robot_1_x))
# print(robot_1_y)
# print(robot_1_z)
# print(robot_2_x)
# print(robot_2_y)
# print(robot_2_z)
# print(robot_3_x)
# print(robot_3_y)
# print(robot_3_z)

# 保存数据
outwb = Workbook()  # 新建文件
wo = outwb.active  # 获取激活文件，确认当前工作表

# 保存一号机器人的数据
careerSheet1 = outwb.create_sheet('robot_1', 0)  # 插入一个sheet叫career，创建当前工作表
careerSheet1.cell(row=1, column=1).value = 'x'
careerSheet1.cell(row=1, column=2).value = 'y'
careerSheet1.cell(row=1, column=3).value = 'z'
for i in range(len(robot_1_x)):
    row_no = i + 2
    careerSheet1.cell(row=row_no, column=1).value = robot_1_x[i]
for i in range(len(robot_1_y)):
    row_no = i + 2
    careerSheet1.cell(row=row_no, column=2).value = robot_1_y[i]
for i in range(len(robot_1_z)):
    row_no = i + 2
    careerSheet1.cell(row=row_no, column=3).value = robot_1_z[i]

# 保存二号机器人的数据
careerSheet2 = outwb.create_sheet('robot_2', 0)  # 插入一个sheet叫career，创建当前工作表
careerSheet2.cell(row=1, column=1).value = 'x'
careerSheet2.cell(row=1, column=2).value = 'y'
careerSheet2.cell(row=1, column=3).value = 'z'
for i in range(len(robot_2_x)):
    row_no = i + 2
    careerSheet2.cell(row=row_no, column=1).value = robot_2_x[i]
for i in range(len(robot_2_y)):
    row_no = i + 2
    careerSheet2.cell(row=row_no, column=2).value = robot_2_y[i]
for i in range(len(robot_2_z)):
    row_no = i + 2
    careerSheet2.cell(row=row_no, column=3).value = robot_2_z[i]

# 保存三号机器人的数据
careerSheet3 = outwb.create_sheet('robot_3', 0)  # 插入一个sheet叫career，创建当前工作表
careerSheet3.cell(row=1, column=1).value = 'x'
careerSheet3.cell(row=1, column=2).value = 'y'
careerSheet3.cell(row=1, column=3).value = 'z'
for i in range(len(robot_3_x)):
    row_no = i + 2
    careerSheet3.cell(row=row_no, column=1).value = robot_3_x[i]
for i in range(len(robot_3_y)):
    row_no = i + 2
    careerSheet3.cell(row=row_no, column=2).value = robot_3_y[i]
for i in range(len(robot_3_z)):
    row_no = i + 2
    careerSheet3.cell(row=row_no, column=3).value = robot_3_z[i]

outwb.save("IPSO-3D-INV-135_to_matlab.xlsx")