# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     horizontal_sampling
   Description :
   Author :       Jamerri
   date：          2022/11/23
-------------------------------------------------
   Change Activity:
                   2022/11/23:
-------------------------------------------------
"""
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# 源位置参数
X_position = 7.35
Y_position = 3.05

# 源定位步数
step_num = 50

# 读取路径
book = load_workbook(filename=r"sampling_xy.xlsx")

# 读取名字为Sheet1的表
sheet1 = book["Sheet1"]

# 用于存储xy坐标
x = []
y = []
row_num = 2
while row_num <= 77:
    # 将表中第一列的2-121行数据写入wind_speed数组中
    x.append(sheet1.cell(row=row_num, column=4).value)
    y.append(sheet1.cell(row=row_num, column=5).value)
    row_num = row_num + 1


# 图像制作
# 画圆线函数
def plot_circle(center=(7.35, 3.05), r=0.5):
    x = np.linspace(center[0] - r, center[0] + r, 5000)
    y1 = np.sqrt(r**2 - (x-center[0])**2) + center[1]
    y2 = -np.sqrt(r**2 - (x-center[0])**2) + center[1]
    plt.plot(x, y1, 'b-', linewidth=1)
    plt.plot(x, y2, 'b-', linewidth=1)


# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

# 设置坐标轴标题
plt.xlabel('X (m)', fontsize=12)  # x轴标题
plt.ylabel('Y (m)', fontsize=12)  # y轴标题

# 设置坐标轴范围
plt.xlim(0, 8)
plt.ylim(0, 4.1)

# 设置坐标轴单位长度一致
ax = plt.gca()
ax.set_aspect(1)

# 绘制折线图，添加数据点，设置点的大小
plt.plot(x, y, '--', c='#f7903d', marker='o', markersize=6, markevery=[0, -1], linewidth=1.75, alpha=0.9)

# 源位置图案
plot_circle((X_position, Y_position), r=0.5)
plt.plot(X_position, Y_position, 'r-', marker='o', markersize=5)

plt.rcParams['figure.figsize'] = (12.0, 6.0)  # 设置figure_size尺寸
plt.savefig(r'horizontal_sampling.svg', bbox_inches='tight', dpi=600)
plt.show()