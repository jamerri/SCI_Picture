#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/3/18 12:52 上午
# @Author : Jamerri
# @FileName: concentration_field.py
# @Email : jamerri@163.com
# @Software: PyCharm


import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import math
from matplotlib.ticker import MultipleLocator, FormatStrFormatter

# 读取路径
book = load_workbook(filename=r"data/浓度场数据-0710/RNV_concentration.xlsx")

# 读取名字为Sheet1的表
sheet = book.get_sheet_by_name("Sheet1")
concentration1 = []

# Imitated 451 Real 451
row_num = 451
while row_num <= 750:
    # 将表中第 i 列的1-100行数据写入数组中
    concentration1.append(sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1


# 读取名字为Sheet2的表
sheet = book.get_sheet_by_name("Sheet2")
concentration2 = []

row_num = 451
while row_num <= 750:
    # 将表中第 i 列的1-100行数据写入数组中
    concentration2.append(sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1

# 读取名字为Sheet1的表
sheet = book.get_sheet_by_name("Sheet3")
concentration3 = []

row_num = 451
while row_num <= 750:
    # 将表中第 i 列的1-100行数据写入数组中
    concentration3.append(sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1


# 图像制作

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=14)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

# X数据
x = np.arange(0, 300, 1)

# 设置坐标轴标题
plt.xlabel('Time (s)', fontsize=14)  # x轴标题
plt.ylabel('Concentration (ppm)', fontsize=14)  # y轴标题

# 设置坐标轴范围
plt.xlim(0, 300)
plt.ylim(0, 400)

# 坐标轴隐藏
ax = plt.gca()
ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax.spines['right'].set_visible(False)  # 隐藏右端脊梁

# 画折线图
l1, = plt.plot(x, concentration1, ls='-', linewidth=1.3, c='#F7903D')
l2, = plt.plot(x, concentration2, ls='-', linewidth=1.3, c='#4D85BD')
l3, = plt.plot(x, concentration3, ls='-', linewidth=1.3, c='#59A95A')

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
# plt.legend(handles=[l1, l2, l3,], labels=['M1', 'M2', 'M3',], loc='best', frameon=True, edgecolor='k')
plt.legend(handles=[l1, l2, l3,], labels=['M1', 'M2', 'M3',], bbox_to_anchor=(0.14, 1.1), loc=2, ncol=3, frameon=False)

# 保存图片
plt.rcParams['figure.figsize'] = (4.0, 3.0)  # 设置figure_size尺寸
plt.savefig(r'pic/RNV_concentration_filed.tif', bbox_inches='tight', dpi=600)
plt.show()
