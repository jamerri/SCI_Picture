#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/3/17 6:13 下午
# @Author : Jamerri
# @FileName: 3D_trajectory.py
# @Email : jamerri@163.com
# @Software: PyCharm


import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.ticker import MultipleLocator, FormatStrFormatter

# 源位置参数
X_position = 7.35
Y_position = 3.05
Z_position = 1.05

# 源定位步数
step_num = 42

# 读取路径
book = load_workbook(filename=r"data/3d_robot_data.xlsx")

# 读取名字为Sheet1的表
sheet1 = book.get_sheet_by_name("Sheet5")

# 读取名字为Sheet2的表
sheet2 = book.get_sheet_by_name("Sheet6")

# 读取名字为Sheet3的表
sheet3 = book.get_sheet_by_name("Sheet7")


# 用于存储1号机器人的数组
robot_1_x = []
robot_1_y = []
robot_1_z = []

row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_1_x.append(sheet1.cell(row=row_num, column=2).value)
    robot_1_y.append(sheet1.cell(row=row_num, column=1).value)
    robot_1_z.append(sheet1.cell(row=row_num, column=3).value)
    row_num = row_num + 1

# 用于存储2号机器人的数组
robot_2_x = []
robot_2_y = []
robot_2_z = []
row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_2_x.append(sheet2.cell(row=row_num, column=2).value)
    robot_2_y.append(sheet2.cell(row=row_num, column=1).value)
    robot_2_z.append(sheet2.cell(row=row_num, column=3).value)
    row_num = row_num + 1

# 用于存储3号机器人的数组
robot_3_x = []
robot_3_y = []
robot_3_z = []
row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_3_x.append(sheet3.cell(row=row_num, column=2).value)
    robot_3_y.append(sheet3.cell(row=row_num, column=1).value)
    robot_3_z.append(sheet3.cell(row=row_num, column=3).value)
    row_num = row_num + 1

# 图像制作


# # 画圆线函数
# def plot_circle(center=(7.35, 3.05), r=0.5):
#     x = np.linspace(center[0] - r, center[0] + r, 5000)
#     y1 = np.sqrt(r**2 - (x-center[0])**2) + center[1]
#     y2 = -np.sqrt(r**2 - (x-center[0])**2) + center[1]
#     plt.plot(x, y1, 'k--')
#     plt.plot(x, y2, 'k--')


# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
# plt.rcParams['xtick.direction'] = 'in'
# plt.rcParams['ytick.direction'] = 'in'
# plt.rcParams['ztick.direction'] = 'in'

# new a figure and set it into 3d
fig = plt.figure()
# ax = fig.gca(projection='3d')
ax = Axes3D(fig)

# 设置坐标轴单位长度一致
# ax.set_aspect(1)

# 设置坐标轴标题
ax.set_xlabel('X (m)', fontsize=12)  # x轴标题
ax.set_ylabel('Y (m)', fontsize=12)  # y轴标题
ax.set_zlabel('Z (m)', fontsize=12)  # y轴标题

# 设置坐标轴范围
ax.set_xlim(0, 8)
ax.set_ylim(0, 4.1)
ax.set_zlim(0, 1.5)

# y轴主刻度设置为1的倍数
axymajorLocator = MultipleLocator(1)
ax.yaxis.set_major_locator(axymajorLocator)

# z轴主刻度设置为0.5的倍数
axzmajorLocator = MultipleLocator(0.5)
ax.zaxis.set_major_locator(axzmajorLocator)

# 绘制折线图，添加数据点，设置点的大小
figure1 = ax.plot(robot_1_x, robot_1_y, robot_1_z, 'b-', marker='o', markersize=3, linewidth=0.5)
figure2 = ax.plot(robot_2_x, robot_2_y, robot_2_z, 'r-', marker='s', markersize=3, linewidth=0.5)
# figure22 = ax.plot(robot_2_x, robot_2_y, 'r--', marker='s', markersize=3, linewidth=0.5)
figure3 = ax.plot(robot_3_x, robot_3_y, robot_3_z, 'k-', marker='^', markersize=3, linewidth=0.5)

# 源位置图案
# plot_circle((X_position, Y_position), r=0.5)
plt.plot(X_position, Y_position, Z_position, c='r', marker='o', markersize=5)

# 起点位置图案
plt.plot(robot_1_x[0], robot_1_y[0], robot_1_z[0], 'k-', marker='o', markersize=4)
plt.plot(robot_2_x[0], robot_2_y[0], robot_2_z[0], 'k-', marker='s', markersize=4)
plt.plot(robot_3_x[0], robot_3_y[0], robot_3_z[0], 'k-', marker='^', markersize=4)

# 终点位置图案
plt.plot(robot_1_x[-1], robot_1_y[-1], robot_1_z[-1], 'k-', marker='o', markersize=4)
plt.plot(robot_2_x[-1], robot_2_y[-1], robot_2_z[-1], 'k-', marker='s', markersize=4)
plt.plot(robot_3_x[-1], robot_3_y[-1], robot_3_z[-1], 'k-', marker='^', markersize=4)

# # 烟羽发现点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
#
# # 烟羽跟踪点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
#
# # 源确认点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)

# plt.rcParams['figure.figsize'] = (6.0, 8.0)  # 设置figure_size尺寸
# plt.savefig(r'pic/trajectory.tiff', bbox_inches='tight', dpi=300)  # 保存图
plt.show()
