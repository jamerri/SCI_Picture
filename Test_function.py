# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     Test_function
   Description :
   Author :       Jamerri
   date：          2022/3/19
-------------------------------------------------
   Change Activity:
                   2022/3/19:
-------------------------------------------------
"""

import os
import shutil
import xlwt
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# 源位置参数
X_position = 7.35
Y_position = 3.05


# 实验组数
Experiment_number = 3


# 数据存储文件夹
dirs = './ddd'
# 文件数量
file_number = 9


# 数据分类函数
def class_files_function(f_num):
    # 获得文件名
    filelist = os.listdir(dirs)
    # print(filelist)

    # 按照实验组数对数据进行分类
    for i in range(int(f_num / 3)):
        path = './ddd/'
        new_dir = os.path.join(path, str(i + 1))
        os.mkdir(new_dir)
        # print(new_dir)
        source_1 = os.path.join(path, filelist[3 * i])
        shutil.move(source_1, new_dir)
        source_2 = os.path.join(path, filelist[3 * i + 1])
        shutil.move(source_2, new_dir)
        source_3 = os.path.join(path, filelist[3 * i + 2])
        shutil.move(source_3, new_dir)


# 对应的文件夹数据合并为Excel
def write_to_excel(f_num):
    for i in range(int(f_num / 3)):
        num = i + 1

        sub_dirs = './ddd/' + str(i + 1)
        sub_filelist = os.listdir(sub_dirs)

        file_path_1 = './ddd/' + str(i + 1) + '/' + sub_filelist[0]
        file_path_2 = './ddd/' + str(i + 1) + '/' + sub_filelist[1]
        file_path_3 = './ddd/' + str(i + 1) + '/' + sub_filelist[2]

        f1 = open(file_path_1, 'r', encoding='utf-8')
        f2 = open(file_path_2, 'r', encoding='utf-8')
        f3 = open(file_path_3, 'r', encoding='utf-8')

        wb = xlwt.Workbook(encoding='utf-8')

        ws1 = wb.add_sheet('1_robot')
        ws2 = wb.add_sheet('2_robot')
        ws3 = wb.add_sheet('3_robot')

        r_txt(f1, ws1, wb, num)
        wb.close()
        r_txt(f2, ws2, wb, num)
        wb.close()
        r_txt(f3, ws3, wb, num)
        wb.close()
    print("Success!!")


# 遍历txt函数
def r_txt(f_n, ws_n, w_b, n):
    row = 1
    col = 0
    k = 1
    for lines in f_n:
        a = lines.split('\t')
        k += 1
        for j in range(len(a)):
            ws_n.write(row, col, a[j])
            col += 1
        row += 1
        col = 0
    w_b.save("./ddd/" + str(n) + "/robots.xlsx")


def plot_trajectory(no_experiment):
    # 读取路径
    fname = './ddd/' + str(no_experiment + 1) + '/robots.xlsx'
    book = load_workbook(fname)

    # 读取名字为Sheet1的表
    sheet1 = book.get_sheet_by_name("1_robot")

    # 读取名字为Sheet2的表
    sheet2 = book.get_sheet_by_name("2_robot")

    # 读取名字为Sheet3的表
    sheet3 = book.get_sheet_by_name("3_robot")

    # 判断此次实验的步数
    step_start = 9
    while step_start <= 100:
        if sheet1.cell(row=step_start, column=2).value == 0:
            break
        else:
            step_start += 1

    step_total = step_start - 9

    # 用于存储1号机器人的数组
    robot_1_x = []
    robot_1_y = []
    row_num = 1
    while row_num <= step_total:
        # 将表中第一列的1-100行数据写入wind_speed数组中
        robot_1_x.append(sheet1.cell(row=row_num, column=2).value)
        robot_1_y.append(sheet1.cell(row=row_num, column=3).value)
        row_num = row_num + 1

    # 用于存储2号机器人的数组
    robot_2_x = []
    robot_2_y = []
    row_num = 1
    while row_num <= step_total:
        # 将表中第一列的1-100行数据写入wind_speed数组中
        robot_2_x.append(sheet2.cell(row=row_num, column=2).value)
        robot_2_y.append(sheet2.cell(row=row_num, column=3).value)
        row_num = row_num + 1

    # 用于存储3号机器人的数组
    robot_3_x = []
    robot_3_y = []
    row_num = 1
    while row_num <= step_total:
        # 将表中第一列的1-100行数据写入wind_speed数组中
        robot_3_x.append(sheet3.cell(row=row_num, column=1).value)
        robot_3_y.append(sheet3.cell(row=row_num, column=2).value)
        row_num = row_num + 1

    # 图像制作

    # 画圆线函数
    def plot_circle(center=(7.35, 3.05), r=0.5):
        x = np.linspace(center[0] - r, center[0] + r, 5000)
        y1 = np.sqrt(r ** 2 - (x - center[0]) ** 2) + center[1]
        y2 = -np.sqrt(r ** 2 - (x - center[0]) ** 2) + center[1]
        plt.plot(x, y1, 'k--')
        plt.plot(x, y2, 'k--')

    # 导入Times New Roman字体
    plt.rc('font', family='Times New Roman', size=12)

    # 设置xtick和ytick的方向：in、out、inout
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    # 设置坐标轴标题
    plt.xlabel('X (m)', fontsize=12)  # x轴标题
    plt.ylabel('Y (m)', fontsize=12)  # y轴标题

    # 设置坐标轴范围
    plt.xlim(0, 8)
    plt.ylim(0, 4.1)

    # 设置坐标轴单位长度一致
    ax = plt.gca()
    ax.set_aspect(1)

    # 绘制折线图，添加数据点，设置点的大小
    plt.plot(robot_1_x, robot_1_y, 'b-', marker='o', markersize=3, linewidth=0.5)
    plt.plot(robot_2_x, robot_2_y, 'r-', marker='s', markersize=3, linewidth=0.5)
    plt.plot(robot_3_x, robot_3_y, 'k-', marker='^', markersize=3, linewidth=0.5)

    # 源位置图案
    plot_circle((X_position, Y_position), r=0.5)
    plt.plot(X_position, Y_position, 'r-', marker='o', markersize=5)

    # 起点位置图案
    plt.plot(robot_1_x[0], robot_1_y[0], 'k-', marker='o', markersize=4)
    plt.plot(robot_2_x[0], robot_2_y[0], 'k-', marker='s', markersize=4)
    plt.plot(robot_3_x[0], robot_3_y[0], 'k-', marker='^', markersize=4)

    # 终点位置图案
    plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
    plt.plot(robot_2_x[-1], robot_2_y[-1], 'k-', marker='s', markersize=4)
    plt.plot(robot_3_x[-1], robot_3_y[-1], 'k-', marker='^', markersize=4)

    # # 烟羽发现点位置图案
    # plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
    #
    # # 烟羽跟踪点位置图案
    # plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
    #
    # # 源确认点位置图案
    # plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)

    plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
    save_path = 'ddd/' + str(no_experiment + 1) + '/2D_trajectory.tiff'
    plt.savefig(save_path, bbox_inches='tight', dpi=300)  # 保存图
    plt.show()


class_files_function(file_number)
write_to_excel(file_number)

No_experiment = 1
for No_experiment in range(Experiment_number + 1):
    plot_trajectory(No_experiment)
    No_experiment += 1
