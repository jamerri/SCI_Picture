#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/3/17 2:12 下午
# @Author : Jamerri
# @FileName: wind.py
# @Email : jamerri@163.com
# @Software: PyCharm


import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import math
from matplotlib.ticker import MultipleLocator, FormatStrFormatter

# 读取路径
book = load_workbook(filename=r"data/浓度场数据-0717/M0-20220717.xlsx")

# 读取名字为Sheet1的表
sheet = book.get_sheet_by_name("Sheet2")

# 用于存储数据的数组
wind_speed = []
wind_direction = []

# Imitated 451(M1-M3) Real 621(M1-M3) (M0)
row_num = 651
while row_num <= 950:
    # 将表中第一列的1-100行数据写入wind_speed数组中
    wind_speed.append(sheet.cell(row=row_num, column=2).value)
    wind_direction.append(sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1

# # 弧度换算
# for i in range(len(wind_direction)):
#     D = (180 + wind_direction[i] * (180 / math.pi))
#     if D < 0:
#         wind_direction[i] = 0
#     else:
#         wind_direction[i] = D
#     i = i + 1

# 弧度换算
for i in range(len(wind_direction)):
    if 0 <= wind_direction[i] <= 3.14159:
        D = 360 - wind_direction[i] * (180 / math.pi)
    else:
        D = abs(wind_direction[i]) * (180 / math.pi)
    wind_direction[i] = D
    i = i + 1

# 图像制作

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=14)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'out'
plt.rcParams['ytick.direction'] = 'out'

# X数据
x = np.arange(0, 300, 1)

fig, ax1 = plt.subplots()
ax1.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax2 = ax1.twinx()  # 共享X轴
ax2.spines['top'].set_visible(False)  # 隐藏上端脊梁

# 设置线条参数
l1, = ax1.plot(x, wind_speed, ls='-', linewidth=1.3, c='#E4392E')
l2, = ax2.plot(x, wind_direction, ls='-', linewidth=1.3, c='#3979F2')

# 设置坐标轴范围
plt.xlim(0, 300)
ax1.set_ylim(0, 1.5)
ax2.set_ylim(0, 360)

# 设置为n的倍数
ax1ymajorLocator = MultipleLocator(0.3)
ax1.yaxis.set_major_locator(ax1ymajorLocator)

# 设置为60的倍数
ax2ymajorLocator = MultipleLocator(60)
ax2.yaxis.set_major_locator(ax2ymajorLocator)

# ax2yminorLocator = MultipleLocator(20)  # 将y轴次刻度标签设置为n的倍数
# ax2.yaxis.set_minor_locator(ax2yminorLocator)

# 设置坐标轴标签
ax1.set_xlabel('Time (s)', fontsize=14)
ax1.set_ylabel('Wind Speed (m/s)', fontsize=14, color='k')
ax2.set_ylabel('Wind Direction ('u'\u00b0'')', fontsize=14, color='k')

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
plt.legend(handles=[l1, l2, ], labels=['wind speed', 'wind direction'], bbox_to_anchor=(0.14, 1.1), loc=2, ncol=2,
           frameon=False)
plt.rcParams['figure.figsize'] = (4.0, 3.0)  # 设置figure_size尺寸
plt.savefig(r'pic/M0_wind_RNV-20220717-2.tif', bbox_inches='tight', dpi=600)  # 保存图片
plt.show()
