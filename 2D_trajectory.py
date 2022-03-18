#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/3/17 7:08 下午
# @Author : Jamerri
# @FileName: 2D_trajectory.py
# @Email : jamerri@163.com
# @Software: PyCharm


import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# 源位置参数
X_position = 7.35
Y_position = 3.05

# 源定位步数
step_num = 50

# 读取路径
book = load_workbook(filename=r"data/robot_data.xlsx")

# 读取名字为Sheet1的表
sheet1 = book.get_sheet_by_name("Sheet1")

# 读取名字为Sheet2的表
sheet2 = book.get_sheet_by_name("Sheet2")

# 读取名字为Sheet3的表
sheet3 = book.get_sheet_by_name("Sheet3")


# 用于存储1号机器人的数组
robot_1_x = []
robot_1_y = []
row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_1_x.append(sheet1.cell(row=row_num, column=1).value)
    robot_1_y.append(sheet1.cell(row=row_num, column=2).value)
    row_num = row_num + 1

# 用于存储2号机器人的数组
robot_2_x = []
robot_2_y = []
row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_2_x.append(sheet2.cell(row=row_num, column=1).value)
    robot_2_y.append(sheet2.cell(row=row_num, column=2).value)
    row_num = row_num + 1

# 用于存储3号机器人的数组
robot_3_x = []
robot_3_y = []
row_num = 1
while row_num <= (step_num + 1):
    # 将表中第一列的1-100行数据写入wind_speed数组中
    robot_3_x.append(sheet3.cell(row=row_num, column=1).value)
    robot_3_y.append(sheet3.cell(row=row_num, column=2).value)
    row_num = row_num + 1

# 图像制作


# 画圆线函数
def plot_circle(center=(7.35, 3.05), r=0.5):
    x = np.linspace(center[0] - r, center[0] + r, 5000)
    y1 = np.sqrt(r**2 - (x-center[0])**2) + center[1]
    y2 = -np.sqrt(r**2 - (x-center[0])**2) + center[1]
    plt.plot(x, y1, 'k--')
    plt.plot(x, y2, 'k--')


# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

# 设置坐标轴标题
plt.xlabel('X (m)', fontsize=12)  # x轴标题
plt.ylabel('Y (m)', fontsize=12)  # y轴标题

# 设置坐标轴范围
plt.xlim(0, 8)
plt.ylim(0, 4.1)

# 设置坐标轴单位长度一致
ax = plt.gca()
ax.set_aspect(1)

# 绘制折线图，添加数据点，设置点的大小
plt.plot(robot_1_x, robot_1_y, 'b-', marker='o', markersize=3, linewidth=0.5)
plt.plot(robot_2_x, robot_2_y, 'r-', marker='s', markersize=3, linewidth=0.5)
plt.plot(robot_3_x, robot_3_y, 'k-', marker='^', markersize=3, linewidth=0.5)

# 源位置图案
plot_circle((X_position, Y_position), r=0.5)
plt.plot(X_position, Y_position, 'r-', marker='o', markersize=5)

# 起点位置图案
plt.plot(robot_1_x[0], robot_1_y[0], 'k-', marker='o', markersize=4)
plt.plot(robot_2_x[0], robot_2_y[0], 'k-', marker='s', markersize=4)
plt.plot(robot_3_x[0], robot_3_y[0], 'k-', marker='^', markersize=4)

# 终点位置图案
plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
plt.plot(robot_2_x[-1], robot_2_y[-1], 'k-', marker='s', markersize=4)
plt.plot(robot_3_x[-1], robot_3_y[-1], 'k-', marker='^', markersize=4)

# # 烟羽发现点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
#
# # 烟羽跟踪点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)
#
# # 源确认点位置图案
# plt.plot(robot_1_x[-1], robot_1_y[-1], 'k-', marker='o', markersize=4)

plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
plt.savefig(r'pic/2D_trajectory.tiff', bbox_inches='tight', dpi=300)  # 保存图
plt.show()

