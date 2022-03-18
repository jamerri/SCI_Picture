#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/3/17 2:12 下午
# @Author : Jamerri
# @FileName: wind.py
# @Email : jamerri@163.com
# @Software: PyCharm


import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import math
from matplotlib.ticker import MultipleLocator, FormatStrFormatter


# 源定位步数
step_num = 50

# 读取路径
book = load_workbook(filename=r"data/concentration_data.xlsx")


# 读取名字为Sheet1的表
sheet = book.get_sheet_by_name("Sheet1")
concentration1 = []

row_num = 9
while row_num <= (step_num + 8):
    # 将表中第 i 列的1-100行数据写入数组中
    concentration1.append(sheet.cell(row=row_num, column=10).value)
    row_num = row_num + 1


# 读取名字为Sheet2的表
sheet = book.get_sheet_by_name("Sheet2")
concentration2 = []

row_num = 9
while row_num <= (step_num + 8):
    # 将表中第 i 列的1-100行数据写入数组中
    concentration2.append(sheet.cell(row=row_num, column=10).value)
    row_num = row_num + 1

# 读取名字为Sheet1的表
sheet = book.get_sheet_by_name("Sheet3")
concentration3 = []

row_num = 9
while row_num <= (step_num + 8):
    # 将表中第 i 列的1-100行数据写入数组中
    concentration3.append(sheet.cell(row=row_num, column=10).value)
    row_num = row_num + 1

concentration = []
for i in range(step_num):
    concentration.append((max(concentration1[i], concentration2[i], concentration3[i])) / 4)
    i = i + 1

# 图像制作

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=15)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

# X数据
x = np.arange(0, step_num, 1)
fig, ax = plt.subplots()

# 设置线条参数
#l1, = ax.plot(x, concentration,  'g-')
ax.plot(x, concentration, marker='s', markersize=3, color='k', linewidth='1', linestyle='-')

ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax.spines['right'].set_visible(False)  # 隐藏右端脊梁

# 设置坐标轴范围
plt.xlim(0, step_num)
ax.set_ylim(0, 100)


# 设置Y轴主刻度
axymajorLocator = MultipleLocator(10)
#ax2.yaxis.set_major_locator(axymajorLocator)

# 设置Y轴次刻度
# axyminorLocator = MultipleLocator(5)
# ax.yaxis.set_minor_locator(axyminorLocator)

# 设置坐标轴标签
ax.set_xlabel('Steps', fontsize=15)
ax.set_ylabel('Concentration (ppm)', fontsize=15, color='k')

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
#plt.legend(handles=[l1, l2,], labels=['wind speed', 'wind direction'], bbox_to_anchor=(0.18, 1.1), loc=2, ncol=2, frameon=False)

# 保存图片
plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
plt.savefig(r'pic/concentration.tiff', bbox_inches='tight', dpi=300)
plt.show()
